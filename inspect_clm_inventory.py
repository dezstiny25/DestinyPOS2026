import os
from openpyxl import load_workbook

path = 'CLM Inventory.xlsx'
print('exists', os.path.exists(path))
wb = load_workbook(path, data_only=True)
print('sheets', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print('sheet', name)
    print('headers', [c.value for c in ws[1][:12]])
    print('row2', [c.value for c in ws[2][:12]])
    break
